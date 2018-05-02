'''
Created on 02.05.2018

@author: chou
'''
import openpyxl
import glob
import os


def Get_Batch_Info(Batch):
            

                    os.chdir(r"P:\Production and Physics Communication\Tool Plan\Lesker")

                    for file_name in glob.glob('Lesker tool planning *.xlsm'):
                        Tool_Plan_name_Lesker = file_name
                        

                    print(Tool_Plan_name_Lesker)
                    Plan_Lekser = openpyxl.load_workbook(Tool_Plan_name_Lesker)

    
                    Plan_Lekser_sheet = Plan_Lekser.get_sheet_by_name('Planning Tool')


                    #Specify the batch#

                    for w in range(10000, 25000):
                        if Plan_Lekser_sheet.cell(row=w, column=2).value == Batch :
                            Batch_Location_Index = w
                    
                    CW = Plan_Lekser_sheet.cell(row=Batch_Location_Index-1, column=2).value
                    Batch = Plan_Lekser_sheet.cell(row=Batch_Location_Index, column=2).value
                    Project = Plan_Lekser_sheet.cell(row=Batch_Location_Index+1, column=2).value
                    Aim = Plan_Lekser_sheet.cell(row=Batch_Location_Index-2, column=4).value          
                    Substrate_Number = Plan_Lekser_sheet.cell(row=Batch_Location_Index-1, column=10).value 
            
                    Substrate_list = []
                    for s in range (0, Substrate_Number):
                        Substrate_list.append(Plan_Lekser_sheet.cell(row=Batch_Location_Index+1+2*s, column=9).value)
            
            
                    Layer_count = 0
                    while Plan_Lekser_sheet.cell(row=Batch_Location_Index-1+Layer_count, column=4).value != None:
                        Layer_count = Layer_count + 1
                    
        
                    Architecture = []
                    for A in range (0,Layer_count):
                        for l in range (0,6):
                            Architecture.append(Plan_Lekser_sheet.cell(row=Batch_Location_Index-1+A, column=3+l).value)
                
            
                
                    Batch_Info= dict({'Calendar_Week' : CW, 'Batch_Number' : Batch, 'Project_Name' : Project, 'Aim' : Aim, 'Substrate_Number': Substrate_Number, 'Substrate_List' : Substrate_list, 'Layer_Number' : Layer_count, 'Architecture' : Architecture})
                
                    return Batch_Info