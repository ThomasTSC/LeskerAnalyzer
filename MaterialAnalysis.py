# -*- coding: utf-8 -*-
"""
Created on Wed Apr 25 18:56:24 2018

@author: chou
"""





import os
import Get_Batch_Info
import glob
import seaborn
import matplotlib.pyplot as plt
import fnmatch
import openpyxl
import xlwings
import File_Manipulation
import Layer_Analysis
import Log_File_Info 
import numpy
import Count_Layer_Number

class MaterialAnalysis:
    
    def __init__(self, Batch_Number):
        
        self.Batch = Batch_Number
        self.Lesker_Folder_Path = 'P:\Forschungs-Projekte\OLED-measurements Bruchsal\B 2018 Lesker'
        self.Layer_Number = Count_Layer_Number.Count_Layer_Number(Batch_Number)




    
    def Load_Fabrication_Form(self):
        
        os.chdir(self.Lesker_Folder_Path)
    

        for file_name in os.listdir(self.Lesker_Folder_Path):
            if fnmatch.fnmatch(file_name, '* Batch %d' %self.Batch):
                Batch_File_Path = ('P:\Forschungs-Projekte\OLED-measurements Bruchsal\B 2018 Lesker\%s' %file_name)
                os.chdir(Batch_File_Path)
            
        for file_name in glob.glob('*-sheet.xlsx'):
           Fabrication_Form_Path = (Batch_File_Path+'\%s' %file_name)
           
        
        Lesker_FF = openpyxl.load_workbook(Fabrication_Form_Path)
        
        Lesker_FF_Sheet =  Lesker_FF.get_sheet_by_name('fabSheet')
        
        return Lesker_FF_Sheet
        
        
        
    
    def Info_per_Layer(self):
        
        Lesker_FF_Sheet = MaterialAnalysis(self.Batch).Load_Fabrication_Form()
        
        Material_per_Layer = []
        ExpID_per_Layer = []
        Source_per_Layer = []

        i= 0

        
        while Lesker_FF_Sheet.cell(row=80+i, column=4).value != None:
            Material_per_Layer.append(Lesker_FF_Sheet.cell(row=80+i, column=4).value)
            ExpID_per_Layer.append(Lesker_FF_Sheet.cell(row=80+i, column=5).value)
            Source_per_Layer.append(Lesker_FF_Sheet.cell(row=80+i, column=8).value)

            i = i+1
    
    



        Source_per_Layer, Material_per_Layer, ExpID_per_Layer = zip(*sorted(zip(Source_per_Layer, Material_per_Layer, ExpID_per_Layer)))
        
        #print (Source_per_Layer, Material_per_Layer, ExpID_per_Layer)
        
        Source_per_Layer = list(dict.fromkeys(Source_per_Layer).keys())
        Material_per_Layer = list(dict.fromkeys(Material_per_Layer).keys())
        ExpID_per_Layer = list(dict.fromkeys(ExpID_per_Layer).keys())

        del Source_per_Layer[0]
        del Material_per_Layer[0]
        del ExpID_per_Layer[0]
        
        Info_per_Layer = {'Source':Source_per_Layer, 'Material':Material_per_Layer, 'ExpID': ExpID_per_Layer}
        
        print (Info_per_Layer)
        
        return Info_per_Layer
    
    
    
    def Temperature_Plot(self):
        
        for Layer_Order in range (0, self.Layer_Number):
            Layer_Data = File_Manipulation.File_Manipulation(self.Batch, Layer_Order).Data_Classification()
            Sensor_Info=Layer_Analysis.Layer_Analysis(self.Batch).Corresponding_Sensor(Layer_Order)
            Time = Layer_Analysis.Layer_Analysis(self.Batch).TimeDuration(Layer_Order)
     
            
            if Layer_Order == self.Layer_Number-1:
                Sensor_Info[1][0]= 6 
            
            
            
            #Time v.s. Rate Plot#
            if len(Sensor_Info[1]) >1:
                plt.figure()
                for k in range (len(Sensor_Info[1])):
                    Reg_Plot = seaborn.regplot(x=numpy.array(Time), y=Layer_Data['OLED_%d' %(int((list(Sensor_Info[0].values())[0])[k]))]  )
            
            else:
                plt.figure()
                Reg_Plot = seaborn.regplot(x=numpy.array(Time), y=Layer_Data['OLED_%d' %Sensor_Info[1][0]])
                
            plt.title('Layer_%d' %(Layer_Order+1) )
            plt.ylabel('Temp (degree)')
            plt.xlabel('Time (s)')
            #plt.savefig(Log_File_Info.Log_File_Info(self.Batch).Log_File_Info()['Batch_File_Path']+"\ "+"Layer_%d" %(Layer_Order+1))
            
    
    
    
    
    def Data_Collection(self):
        
        pass



#Test#
#MaterialAnalysis(661).Load_Fabrication_Form()
MaterialAnalysis(658).Info_per_Layer()
#MaterialAnalysis(661).Temperature_Plot()